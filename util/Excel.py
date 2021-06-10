#!/usr/bin/python
# -*- coding: utf-8 -*-
# @Time    : 2020/10/16 14:11
# @Author  : hongjun.xiao
# @File    : Excel.py
# @system  : WenJiang


from openpyxl import load_workbook
import os
from util import ProjectPath

#读取Excel的方法
def ReadExcel():
    wb=load_workbook(ProjectPath.PtPath("/casedata/GovernmentCaseData.xlsx"))
    sheet=wb["Sheet1"]
    allvalue=[]
    for row in list(sheet.rows)[1:sheet.max_row]:
        value=[]
        for column in range(0,sheet.max_column):
            value.append(row[column].value)
        if value[2]=="是":
            allvalue.append(value)
        else:
            continue
    return allvalue


def ReadExcel_new():
    wb=load_workbook(ProjectPath.PtPath("/casedata/GovernmentCaseData.xlsx"))
    sheet=wb["Sheet2"]
    #处理合并到的单元格，返回每个单元格所占了几行
    merged_list=list(str(sheet.merged_cells))
    print(a)
    rownus = []
    rownus_original=[]
    n=0
    for A_list in merged_list:
        n=n+1
        if A_list=="A":
            #ans=a.index(A_list)获取列表元素的索引
            rownus_original.append(merged_list[n])
            if len(rownus_original)==2:
                rownus.append(rownus_original)
                rownus_original=[]
            else:
                continue
        else:
            continue
    print(rownus)
    #获取excel中基础字段的最初值
    general_value=[]
    for row in list(sheet.rows)[1:sheet.max_row]:
        normal_value=[]
        for column in range(0,6):
            normal_value.append(row[column].value)
        general_value.append(normal_value)
    # 去除掉获取到的值中的空列表
    all_general_value=[]
    for list_element in general_value:
        if list_element[0]!=None:
            all_general_value.append(list_element)
        else:
            continue
    #获取到操作描述以及操作元素
    all_step_value=[]
    for num in rownus:
        step_value_list=[]
        for row1 in list(sheet.rows)[int(num[0])-1:int(num[1])]:
            step_value=[]
            for column1 in range(6,10):
                step_value.append(row1[column1].value)
            step_value_list=step_value_list+step_value
        all_step_value.append(step_value_list)

    #将用例描述和用例步骤拼接起来组成一个新的列表
    all_case_value=[]
    for i in range(len(all_general_value)):
        all_case_value.append(all_general_value[i]+all_step_value[i])

    return all_case_value









def ExecuteTwo():
    wb=load_workbook(ProjectPath.PtPath("/casedata/GovernmentCaseData.xlsx"))
    sheet=wb["Sheet1"]
    allvalue=[]
    for row in list(sheet.rows)[1:sheet.max_row]:
        value=[]
        for column in range(0,sheet.max_column):
            value.append(row[column].value)
        if value[1]=="Fail":
            allvalue.append(value)
        else:
            continue
    return allvalue



def WriteExcel(row,results):
    wb=load_workbook(ProjectPath.PtPath("/casedata/GovernmentCaseData.xlsx"))
    sheet=wb["Sheet1"]
    sheet.cell(row=row,column=2).value=results
    wb.save(ProjectPath.PtPath("/casedata/GovernmentCaseData.xlsx"))



if __name__=="__main__":
    print(ReadExcel_new())




